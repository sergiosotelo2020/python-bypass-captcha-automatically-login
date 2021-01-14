import csv
import chilkat
import sys
import re
import os
import json
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import InvalidArgumentException
from datetime import date
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException 
from selenium.webdriver.common.action_chains import ActionChains
import selenium.webdriver
from PIL import Image
import pytesseract
import time
from datetime import datetime
from openpyxl import load_workbook
import xlrd
import pandas as pd
from xlsxwriter.workbook import Workbook


today = date.today()
# now = datetime.now()
current_time = today.strftime("%m%d%y")
print(current_time)
file_name = input("Enter the excel file name: ")
driver = selenium.webdriver.Chrome()

base_url = 'https://sso.gem.gov.in/ARXSSO/oauth/login'

driver.get(base_url)
time.sleep(5)
driver.implicitly_wait(5)
element = driver.find_element_by_id("captcha1")
location = element.location
size = element.size
driver.save_screenshot("pageImage.png")

# crop image
x = location['x']
y = location['y']
width = location['x']+size['width']
height = location['y']+size['height']
im = Image.open('pageImage.png')
im = im.crop((int(x), int(y), int(width), int(height)))
im.save('element.png')
pytesseract.pytesseract.tesseract_cmd = r'C:/Users/Master/AppData/Local/tesseract.exe'
text = pytesseract.image_to_string(Image.open("element.png"))
print(text)
# element=driver.find_element_by_xpath('//span[@class="last-page"]')
ComapnyName = "Appex Corporate Solutions"
GemUsername = "Appex_corp22"
GemPassword = "Mybox@005"
print("----------start login---------------")
user_id = driver.find_element_by_id('loginid')
user_id.send_keys(GemUsername)
captcha_text = driver.find_element_by_id('captcha_math')
captcha_text.send_keys(text)
print("--------Please insert key manually--------------")
# i=driver.find_elements_by_css_selector('button.pagination__btn')[2].text

# img_data = driver.find_element_by_id('captcha1').get_attribute('src')
time.sleep(5)
submit_button=driver.find_elements_by_css_selector("button.btn-nov")[0]
submit_button.click()

time.sleep(1)
password = driver.find_element_by_id('password')
password.send_keys(GemPassword)
time.sleep(2)
submit_button2 = driver.find_element_by_xpath('//button[@type="submit"]')
submit_button2.click()
driver.maximize_window()

#Read excel file
time.sleep(8)

edit_urls = []
statuss = []

print('--------------read excel file-------------------')
workbook = load_workbook(filename=file_name)
# sheet_name = workbook.sheetnames
# print(sheet_name)
sheet = workbook.active
urls = []
x = 0
for i in sheet['F']: 
    url = i.value  
    if (url == None):
        break
    urls.append(url)
urls.remove("Direct Url")
x = len(urls)
print(x)
# try:
#     close_button = driver.find_element_by_id("flox-chat-close")
#     close_button.click()
# except NoSuchElementException:
#     print("don't exist such element")
loc = file_name
 
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)

# sheet1 = xlrd.open_workbook("Web2Pair.xlsx").sheet_by_index(0) 

key = 1
for url in urls:
    
    driver.get(url)
    print('-----------redirect to edit page---------------------')
    time.sleep(10)
    if driver.find_element_by_id("flox-chat-close"):
        close_button = driver.find_element_by_id("flox-chat-close")
        try:
            close_button.click()
            
        except:
            print('here')
    i = 0
    
    while i < 3:
        try:
            driver.find_element_by_class_name("popup-footer")
            popup = driver.find_element_by_class_name("popup-footer")
            popup_button = popup.find_elements_by_css_selector("*")[0]
            popup_button.click()
            time.sleep(10)
            j = 0
            while j < 3:
                try:
                    driver.find_element_by_class_name("img-guidelines")
                    print('go')
                    print(sheet.row_values(key))
                    
                    rows = sheet.row_values(key)
                    
                    ProductCatalogID = rows[0]
                    print(ProductCatalogID)
                    Model = rows[1]
                    Category = rows[2]
                    Brand = rows[3]
                    url = rows[4]
                    Direct_url = rows[5]
                    Authorization_no = rows[6]
                    Authorization__agency = rows[7]
                    Authorization_date = rows[8]
                    From = rows[9]
                    To = rows[10]
                    Country_of_origin = rows[11]
                    sku = str(rows[12])
                    print(sku)
                    hsn = int(rows[13])
                    print(hsn)
                    mrp = int(rows[14])
                    Offer_price = str(rows[15])
                    print(Offer_price)
                    try:
                        Pincodes = int(rows[16])
                    except:
                        Pincodes = ''
                    Disticts = rows[17]
                    print(Disticts)
                    State = rows[18]
                    Current_stock = int(rows[19])
                    mqpc = int(rows[20])
                    print(mqpc)
                    lead_time = int(rows[21])
                    edit_url = rows[22]
                    status = rows[23]
                    print(status)
                    workbook.close() 

                    # third_part = driver.find_element_by_class_name("stock-section-fieldset")
                    try:
                        country = driver.find_elements_by_css_selector("input.input-xs")[0]
                        print(country)
                        country.send_keys(Country_of_origin)
                        time.sleep(1)
                        driver.find_element_by_class_name("ui-select-choices-row-inner").click()
                    except:
                        print('no country')
                    try:
                        sku_value = driver.find_element_by_xpath('//div[@class="row"]/div[@class="col-sm-6"]/input[@type="text"]')
                        print(sku_value)
                        # actions = ActionChains(driver)
                        # actions.move_to_element(sku_value).perform()
                        # sku_value.click()
                        sku_value.send_keys(sku)
                    except:
                        print('no sku')
                    time.sleep(1)
                    
                    try:
                        hsn_value = driver.find_element_by_xpath('//div[@class="row"]/div[@class="col-sm-6 wsp-tool-tip-wrap"]/input[@type="text"]')
                        print(hsn_value)
                        hsn_value.send_keys(hsn)
                        
                    except:
                        print("sorry1")
                    time.sleep(1)
                    
                    try:
                        mrp_value = driver.find_element_by_xpath('//div[@class="row ng-scope"]/div[@class="col-sm-6 tool-tip-wrap"]/input[@type="number"]')
                        print(mrp_value)
                        mrp_value.send_keys(mrp)
                        
                    except:
                        print('sorry2')
                    time.sleep(1)

                    try:
                        offer_price_value = driver.find_element_by_xpath('//div[@class="row"]/div[@class="col-sm-5 wsp-tool-tip-wrap"]/input[@type="number"]')
                        offer_price_value.send_keys(Offer_price)
                        print(offer_price_value)
                    except:
                        print('sorry3')
                    
                    time.sleep(1)
                    try:
                        disticts_value = driver.find_elements_by_xpath('//div[@class="ui-select-container ui-select-multiple ui-select-bootstrap dropdown form-control ng-pristine ng-untouched ng-valid ng-scope ng-empty"]/div/input[@type="search"]')
                        print(disticts_value[0])
                        disticts_value[0].send_keys(Disticts)
                    except:
                        print('no disticts')
                        
                    time.sleep(1)
                    if Disticts == '':
                        print("N/A")
                    else:
                        try:
                            driver.find_elements_by_xpath('//div[@class="ui-select-choices-row ng-scope active"]/span[@class="ui-select-choices-row-inner"]')[1].click()
                        except:
                            print('here')
                    time.sleep(1)
                    print("1")
                    try:
                        pincode = driver.find_element_by_xpath('//div[@class="panel-body"]/div[@class="ui-select-container ui-select-multiple ui-select-bootstrap dropdown form-control ng-pristine ng-untouched ng-valid ng-scope ng-empty"]/div/input[@type="search"]')
                        print("2")
                        pincode.send_keys(Pincodes)
                    except:
                        print('no pincode')
                    time.sleep(1)
                    print('3')
                    if Pincodes == '':
                        print("N/A")
                    else:
                        try:
                            driver.find_element_by_xpath('//div[@class="ui-select-choices-row ng-scope active"]/span[@class="ui-select-choices-row-inner"]').click()
                        except:
                            print('here')
                    time.sleep(2)
                    try:
                        stock = driver.find_elements_by_xpath('//div[@class="row"]/div[@class="col-sm-6 tool-tip-wrap"]/input[@type="number"]')
                        current_stock_input = stock[0]
                        current_stock_input.send_keys(Current_stock)
                        mqpc_input = stock[1]
                        mqpc_input.send_keys(mqpc)
                        lead_time_input = stock[2]
                        lead_time_input.send_keys(lead_time)
                    except:
                        print('no stock')
                    if State == '':
                        print('sorry4')
                    else:
                        states = driver.find_elements_by_xpath('//td/span/input[@type="checkbox"]')
                        for state in states:
                            try:
                                state.click()
                            
        
                            except:
                                print('sorry5')
                    time.sleep(5)
                    print("save")
                    try:
                        driver.find_element_by_xpath('//div[@class="col-sm-6"]/button[@class="button make-model-submit ng-scope ng-isolate-scope"]').click()
                        time.sleep(5)
                        driver.find_element_by_xpath('//div/a[@class="button success-button"]').click()
                    except:
                        print('here')

                    time.sleep(5)
                    try:
                        driver.switch_to.window(driver.window_handles[1])
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        print("end")
                    except:
                        print('no button')
                    time.sleep(1)
                    try:
                        agree = driver.find_element_by_xpath('//div[@class="text-center agree ng-binding ng-scope"]/input[@type="checkbox"]')
                        agree.click()
                    except:
                        print('here')
                    time.sleep(2)
                    edit_url = driver.current_url
                    print(edit_url)
                    edit_urls.append(edit_url)
                    try:
                        driver.find_element_by_xpath('//input[@value="PUBLISH"]').click()
                    except:
                        print('no publish')
                    time.sleep(1)
                    try:
                        driver.find_element_by_xpath('//div/div[@class="btn btn-primary"]').click()
                        time.sleep(5)
                        statu = driver.find_elements_by_xpath('//tr/td[@id="status"]')[0]
                        print('status')
                    except:
                        statu = 'Already exits'
                        print('status')
                    statuss.append(statu)

                    time.sleep(10)
                    print("complete")
                    break

                except:
                    j += 1
                    print('refresh1')
                    driver.refresh()
                    time.sleep(10)

            break
        except:
            print('no popup')
            driver.refresh()
            time.sleep(15)
            i += 1
            continue

    time.sleep(5)
    print("done" + str(key))
    key += 1
    continue  

wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
# sheet1 = xlrd.open_workbook("Web2Pair.xlsx").sheet_by_index(0) 
tt = open("T.csv", 'w', newline="")
col = csv.writer(tt) 

for row in range(sheet.nrows): 
    col.writerow(sheet.row_values(row))  
# df = pd.DataFrame(pd.read_csv("T.csv")) 
# df
tt.close()
csvv = chilkat.CkCsv()
csvv.put_HasColumnNames(True)
time.sleep(1)

success = csvv.LoadFile("T.csv")
if (success != True):
    print(csvv.lastErrorText())
    sys.exit()
k = 0
ee = len(edit_urls)
print('urls:' + str(ee))
ss = len(statuss)
print('status:' + str(ss))
if ee < ss:
    ee = ss

while k < ee:
    success = csvv.SetCell(k,22,edit_urls[k])
    success = csvv.SaveFile("V.csv")
    success = csvv.SetCell(k,23,statuss[k])
    success = csvv.SaveFile("V.csv")
    k += 1


if (success != True):
    print(csvv.lastErrorText())
time.sleep(1)
csvfile = "V.csv"

workbook = Workbook(file_name)
worksheet = workbook.add_worksheet()
with open(csvfile, 'rt', encoding='utf8') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            worksheet.write(r, c, col)
workbook.close()  

time.sleep(2)
os.remove("T.csv")
os.remove("V.csv")

print("done")