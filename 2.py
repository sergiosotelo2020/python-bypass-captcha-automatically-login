import csv
import sys
import re
import os
import json
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import InvalidArgumentException
from datetime import date
from selenium.webdriver.support.ui import WebDriverWait
import selenium.webdriver
from io import BytesIO
import lxml.html
from PIL import Image
import requests
import pytesseract
import time
from datetime import datetime
import urllib.request
import pandas as pd

from openpyxl import load_workbook


workbook = load_workbook(filename="Web2Pair.xlsx")
# sheet_name = workbook.sheetnames
# print(sheet_name)
sheet = workbook.active
urls = []
x=0
for i in sheet['F']: 
    url = i.value  
    if (url == None):
        break
    urls.append(url)
urls.remove("Direct Url")
x = len(urls)
print(x)

data = pd.ExcelFile("Web2Pair.xlsx")
sheet_names = data.sheet_names
sheet_name = sheet_names[0]
print(sheet_name)

df = data.parse(sheet_name)
df.info 

for row in range(2, x + 2):
    ProductCatalogID = sheet['A' + str(row)].value
    print(ProductCatalogID)

import xlrd
 
loc = ("Web2Pair.xlsx")
 
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
print(sheet.row_values(1))